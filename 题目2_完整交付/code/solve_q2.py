"""Reproducible solver for 2025 CUMCM Problem B, Question 2.

The script solves only the SiC part of Question 2.  It reads the two copied
workbooks in ``../data`` and writes all numerical products to ``../results``.
The raw workbooks are never modified.

Model summary
-------------
* A Lorentz--Drude complex dielectric function represents the dispersive SiC
  epitaxial layer and substrate.
* The Question-1 two-beam Fresnel model is evaluated for both s and p
  polarisations and averaged for unpolarised incident light.
* The thickness is shared by the 10 and 15 degree spectra.  The two spectra
  have separate free-carrier parameters because their doping levels need not
  be identical in the effective model.
* A bounded affine calibration ``observed = scale * physical + offset`` is
  eliminated analytically at every objective evaluation.  This accounts for
  the fact that the supplied reflectance percentages include instrument
  normalisation; it does not change the phase/thickness model.
* Differential evolution explores the bounded parameter box on a fixed
  subsample; L-BFGS-B then refines the best candidate on all points.

The numerical values of the Lorentz--Drude background are modelling
assumptions (standard 4H-SiC infrared constants), not quantities stated in
the contest attachment.  They are recorded in ``q2_fit.json``.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import platform
import sys
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Any, Callable, Iterable

import numpy as np
from openpyxl import load_workbook
from scipy.optimize import differential_evolution, minimize, minimize_scalar
from scipy.signal import find_peaks, savgol_filter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DATA_DIR = ROOT / "data"
RESULT_DIR = ROOT / "results"

SIC_CONSTANTS = {
    "eps_inf": 6.56,
    "omega_L_cm_inv": 970.0,
    "omega_T_cm_inv": 798.0,
}

ANGLES_DEG = {1: 10.0, 2: 15.0}
PARAMETER_NAMES = (
    "d_um",
    "gamma_L_cm_inv",
    "gamma_T_cm_inv",
    "omega_p_layer_cm_inv",
    "gamma_p_layer_cm_inv",
    "omega_p_substrate_cm_inv",
    "gamma_p_substrate_cm_inv",
)

# Broad positive bounds.  The thickness box covers the period-based initial
# estimates with margin and avoids selecting a neighbouring phase branch.
PARAMETER_BOUNDS = (
    (3.0, 12.0),
    (0.1, 1500.0),
    (0.1, 1500.0),
    (1.0, 5000.0),
    (0.1, 2000.0),
    (1.0, 5000.0),
    (0.1, 2000.0),
)


@dataclass(frozen=True)
class Spectrum:
    """Cleaned spectrum plus immutable-source audit metadata."""

    attachment: int
    angle_deg: float
    sigma_cm_inv: np.ndarray
    reflectance_percent: np.ndarray
    source_path: str
    headers: tuple[str, ...]
    source_rows: int
    removed_first_zero: bool


def json_ready(value: Any) -> Any:
    """Convert NumPy/scientific values to JSON-compatible values."""

    if isinstance(value, dict):
        return {str(key): json_ready(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [json_ready(item) for item in value]
    if isinstance(value, np.ndarray):
        return [json_ready(item) for item in value.tolist()]
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return float(value)
    if isinstance(value, (np.bool_,)):
        return bool(value)
    if isinstance(value, float) and not np.isfinite(value):
        return None
    return value


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(json_ready(payload), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def read_workbook(path: Path) -> tuple[np.ndarray, tuple[str, ...]]:
    """Read the first two columns of the first sheet without changing it."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = tuple(str(item) for item in header_row[:2])
    rows = [
        (row[0], row[1])
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row[0] is not None and row[1] is not None
    ]
    workbook.close()
    values = np.asarray(rows, dtype=float)
    if values.ndim != 2 or values.shape[1] != 2 or values.shape[0] < 10:
        raise ValueError(f"{path} does not contain a usable two-column spectrum")
    if not np.all(np.isfinite(values)):
        raise ValueError(f"{path} contains non-finite numeric cells")
    return values, headers


def load_spectrum(attachment: int, data_dir: Path = DEFAULT_DATA_DIR) -> Spectrum:
    """Load one attachment and remove only an exactly-zero first endpoint."""

    path = data_dir / f"附件{attachment}.xlsx"
    values, headers = read_workbook(path)
    sigma = values[:, 0]
    reflectance = values[:, 1]
    removed = bool(len(reflectance) and np.isclose(reflectance[0], 0.0, atol=1e-12))
    if removed:
        sigma = sigma[1:]
        reflectance = reflectance[1:]
    if np.any(np.diff(sigma) <= 0):
        raise ValueError(f"{path} has non-increasing wavenumbers after cleaning")
    return Spectrum(
        attachment=attachment,
        angle_deg=ANGLES_DEG[attachment],
        sigma_cm_inv=sigma,
        reflectance_percent=reflectance,
        source_path=str(path),
        headers=headers,
        source_rows=int(values.shape[0]),
        removed_first_zero=removed,
    )


def audit_spectra(spectra: tuple[Spectrum, Spectrum]) -> dict[str, Any]:
    """Write a compact, byte-traceable input audit."""

    entries: list[dict[str, Any]] = []
    for spectrum in spectra:
        raw_path = Path(spectrum.source_path)
        raw, headers = read_workbook(raw_path)
        spacing = np.diff(raw[:, 0])
        entries.append(
            {
                "attachment": spectrum.attachment,
                "angle_deg": spectrum.angle_deg,
                "source_file": str(raw_path),
                "source_sha256": sha256(raw_path),
                "sheet": "active_sheet",
                "headers": headers,
                "source_rows": int(raw.shape[0]),
                "fitting_rows": int(spectrum.sigma_cm_inv.size),
                "removed_first_zero": spectrum.removed_first_zero,
                "raw_wavenumber_cm_inv": {
                    "min": float(raw[:, 0].min()),
                    "max": float(raw[:, 0].max()),
                    "median_step": float(np.median(spacing)),
                    "min_step": float(spacing.min()),
                    "max_step": float(spacing.max()),
                    "duplicate_count": int(np.sum(np.isclose(spacing, 0.0))),
                },
                "raw_reflectance_percent": {
                    "min": float(raw[:, 1].min()),
                    "max": float(raw[:, 1].max()),
                    "mean": float(raw[:, 1].mean()),
                    "above_100_count": int(np.sum(raw[:, 1] > 100.0)),
                    "zero_row_indices_zero_based": np.where(
                        np.isclose(raw[:, 1], 0.0, atol=1e-12)
                    )[0],
                },
                "fitting_range_cm_inv": [
                    float(spectrum.sigma_cm_inv.min()),
                    float(spectrum.sigma_cm_inv.max()),
                ],
            }
        )
    payload = {
        "question": "Question 2 only",
        "cleaning_rule": "Remove only the first physical data row when its reflectance is exactly 0%; keep the raw workbook unchanged.",
        "attachments": entries,
    }
    write_json(RESULT_DIR / "data_audit.json", payload)
    return payload


def period_initialisation(spectrum: Spectrum) -> dict[str, Any]:
    """Use Question-1 adjacent-peak spacing as a transparent initial value."""

    mask = (spectrum.sigma_cm_inv >= 1300.0) & (spectrum.sigma_cm_inv <= 3500.0)
    sigma = spectrum.sigma_cm_inv[mask]
    response = spectrum.reflectance_percent[mask]
    window = min(101, response.size - (1 - response.size % 2))
    if window % 2 == 0:
        window -= 1
    smooth = savgol_filter(response, window, 3)
    peaks, props = find_peaks(smooth, prominence=0.3, distance=250)
    peak_sigma = sigma[peaks]
    spacing = np.diff(peak_sigma)
    if spacing.size == 0:
        raise ValueError(f"No adjacent peaks detected for attachment {spectrum.attachment}")
    median_spacing = float(np.median(spacing))
    n_initial = 2.55
    optical_factor = np.sqrt(
        n_initial**2 - np.sin(np.deg2rad(spectrum.angle_deg)) ** 2
    )
    d_initial_um = float(1e4 / (2.0 * optical_factor * median_spacing))
    return {
        "attachment": spectrum.attachment,
        "angle_deg": spectrum.angle_deg,
        "detection_band_cm_inv": [1300.0, 3500.0],
        "smoothing": {
            "method": "Savitzky-Golay",
            "window_points": int(window),
            "polynomial_order": 3,
        },
        "peaks_cm_inv": peak_sigma,
        "peak_prominence_percent": props["prominences"],
        "successive_spacing_cm_inv": spacing,
        "median_spacing_cm_inv": median_spacing,
        "n_initial": n_initial,
        "d_initial_um": d_initial_um,
        "role": "Question-1 period estimate used only for initialization and an independent scale check.",
    }


def forward_root(value: np.ndarray | complex) -> np.ndarray:
    """Select the refractive-index branch with non-negative attenuation."""

    root = np.sqrt(np.asarray(value, dtype=np.complex128))
    flip_real = np.real(root) < 0
    root = np.where(flip_real, -root, root)
    # The propagation convention below is exp(+i*delta); conjugating a root
    # with negative imaginary part makes the wave attenuate rather than grow.
    root = np.where(np.imag(root) < 0, np.conj(root), root)
    return root


def forward_cosine(value: np.ndarray | complex) -> np.ndarray:
    """Select the transmitted cosine branch pointing into the next medium."""

    cosine = np.sqrt(np.asarray(value, dtype=np.complex128))
    flip = (np.real(cosine) < 0) | (
        (np.abs(np.real(cosine)) < 1e-14) & (np.imag(cosine) < 0)
    )
    return np.where(flip, -cosine, cosine)


def lorentz_drude_dielectric(
    sigma_cm_inv: np.ndarray,
    gamma_l_cm_inv: float,
    gamma_t_cm_inv: float,
    omega_p_cm_inv: float,
    gamma_p_cm_inv: float,
    constants: dict[str, float] | None = None,
) -> np.ndarray:
    """Complex SiC dielectric function in wavenumber units."""

    c = SIC_CONSTANTS if constants is None else constants
    sigma = np.asarray(sigma_cm_inv, dtype=np.complex128)
    lattice = c["eps_inf"] * (
        (c["omega_L_cm_inv"] ** 2 - sigma**2 - 1j * gamma_l_cm_inv * sigma)
        / (c["omega_T_cm_inv"] ** 2 - sigma**2 - 1j * gamma_t_cm_inv * sigma)
    )
    drude = omega_p_cm_inv**2 / (
        sigma**2 + 1j * gamma_p_cm_inv * sigma
    )
    return lattice - drude


def sic_refractive_index(
    sigma_cm_inv: np.ndarray,
    gamma_l_cm_inv: float,
    gamma_t_cm_inv: float,
    omega_p_cm_inv: float,
    gamma_p_cm_inv: float,
    constants: dict[str, float] | None = None,
) -> np.ndarray:
    return forward_root(
        lorentz_drude_dielectric(
            sigma_cm_inv,
            gamma_l_cm_inv,
            gamma_t_cm_inv,
            omega_p_cm_inv,
            gamma_p_cm_inv,
            constants,
        )
    )


def fresnel_amplitudes(
    n_i: np.ndarray | complex,
    n_j: np.ndarray | complex,
    theta_i_rad: float | np.ndarray,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    """Return s/p r,t amplitudes and the transmitted cosine."""

    n_i_arr = np.asarray(n_i, dtype=np.complex128)
    n_j_arr = np.asarray(n_j, dtype=np.complex128)
    theta_i = np.asarray(theta_i_rad, dtype=np.complex128)
    cos_i = np.cos(theta_i)
    sin_j = n_i_arr * np.sin(theta_i) / n_j_arr
    cos_j = forward_cosine(1.0 - sin_j**2)

    denominator_s = n_i_arr * cos_i + n_j_arr * cos_j
    denominator_p = n_j_arr * cos_i + n_i_arr * cos_j
    r_s = (n_i_arr * cos_i - n_j_arr * cos_j) / denominator_s
    t_s = 2.0 * n_i_arr * cos_i / denominator_s
    r_p = (n_j_arr * cos_i - n_i_arr * cos_j) / denominator_p
    t_p = 2.0 * n_i_arr * cos_i / denominator_p
    return r_s, r_p, t_s, t_p, cos_j


def sic_two_beam_percent(
    sigma_cm_inv: np.ndarray,
    parameters: Iterable[float],
    angle_deg: float,
    constants: dict[str, float] | None = None,
) -> np.ndarray:
    """Question-1 two-beam unpolarised reflectance, in percent."""

    (
        d_um,
        gamma_l,
        gamma_t,
        omega_p_layer,
        gamma_p_layer,
        omega_p_substrate,
        gamma_p_substrate,
    ) = np.asarray(tuple(parameters), dtype=float)
    theta0 = np.deg2rad(float(angle_deg))
    n_layer = sic_refractive_index(
        sigma_cm_inv,
        gamma_l,
        gamma_t,
        omega_p_layer,
        gamma_p_layer,
        constants,
    )
    n_substrate = sic_refractive_index(
        sigma_cm_inv,
        gamma_l,
        gamma_t,
        omega_p_substrate,
        gamma_p_substrate,
        constants,
    )
    r01s, r01p, t01s, t01p, cos1 = fresnel_amplitudes(1.0, n_layer, theta0)
    theta1 = np.arccos(cos1)
    r12s, r12p, _, _, _ = fresnel_amplitudes(n_layer, n_substrate, theta1)
    _, _, t10s, t10p, _ = fresnel_amplitudes(n_layer, 1.0, theta1)

    d_cm = float(d_um) * 1e-4
    delta = 4.0 * np.pi * np.asarray(sigma_cm_inv, dtype=np.complex128) * d_cm * n_layer * cos1
    propagation = np.exp(1j * delta)
    amp_s = r01s + t01s * r12s * t10s * propagation
    amp_p = r01p + t01p * r12p * t10p * propagation
    return np.real((np.abs(amp_s) ** 2 + np.abs(amp_p) ** 2) * 50.0)


def affine_calibration(
    physical_percent: np.ndarray,
    observed_percent: np.ndarray,
    use_calibration: bool = True,
) -> tuple[float, float, np.ndarray]:
    """Eliminate a bounded per-angle scale and offset by least squares."""

    physical = np.asarray(physical_percent, dtype=float)
    observed = np.asarray(observed_percent, dtype=float)
    if not use_calibration:
        return 1.0, 0.0, physical
    design = np.column_stack((physical, np.ones_like(physical)))
    scale, _ = np.linalg.lstsq(design, observed, rcond=None)[0]
    scale = float(np.clip(scale, 0.25, 1.75))
    # With scale fixed/clipped, the best offset is the mean residual.
    offset = float(np.clip(np.mean(observed - scale * physical), -30.0, 30.0))
    prediction = scale * physical + offset
    return scale, offset, prediction


def fit_metrics(observed: np.ndarray, prediction: np.ndarray) -> dict[str, float]:
    residual = np.asarray(prediction, dtype=float) - np.asarray(observed, dtype=float)
    sse = float(np.sum(residual**2))
    sst = float(np.sum((observed - np.mean(observed)) ** 2))
    return {
        "sse_percent_squared": sse,
        "rmse_percent": float(np.sqrt(np.mean(residual**2))),
        "mae_percent": float(np.mean(np.abs(residual))),
        "max_abs_residual_percent": float(np.max(np.abs(residual))),
        "mean_residual_percent": float(np.mean(residual)),
        "residual_q50_abs_percent": float(np.quantile(np.abs(residual), 0.50)),
        "residual_q95_abs_percent": float(np.quantile(np.abs(residual), 0.95)),
        "r2": float(1.0 - sse / sst) if sst > 0 else float("nan"),
    }


def decode(unit: np.ndarray, bounds: tuple[tuple[float, float], ...]) -> np.ndarray:
    lower = np.asarray([item[0] for item in bounds], dtype=float)
    upper = np.asarray([item[1] for item in bounds], dtype=float)
    return lower + np.clip(np.asarray(unit, dtype=float), 0.0, 1.0) * (upper - lower)


def encode(parameters: Iterable[float], bounds: tuple[tuple[float, float], ...]) -> np.ndarray:
    values = np.asarray(tuple(parameters), dtype=float)
    lower = np.asarray([item[0] for item in bounds], dtype=float)
    upper = np.asarray([item[1] for item in bounds], dtype=float)
    return np.clip((values - lower) / (upper - lower), 0.0, 1.0)


def evaluate(
    parameters: np.ndarray,
    spectra: tuple[Spectrum, Spectrum],
    use_calibration: bool = True,
    constants: dict[str, float] | None = None,
) -> tuple[float, list[dict[str, Any]]]:
    """Evaluate both angles and return SSE plus arrays for diagnostics."""

    total_sse = 0.0
    evaluations: list[dict[str, Any]] = []
    for spectrum in spectra:
        physical = sic_two_beam_percent(
            spectrum.sigma_cm_inv, parameters, spectrum.angle_deg, constants
        )
        if not np.all(np.isfinite(physical)) or np.any(physical < -1e-8):
            return float("inf"), []
        scale, offset, prediction = affine_calibration(
            physical, spectrum.reflectance_percent, use_calibration
        )
        metrics = fit_metrics(spectrum.reflectance_percent, prediction)
        total_sse += metrics["sse_percent_squared"]
        evaluations.append(
            {
                "attachment": spectrum.attachment,
                "angle_deg": spectrum.angle_deg,
                "scale": scale,
                "offset_percent": offset,
                "physical_percent": physical,
                "prediction_percent": prediction,
                "residual_percent": prediction - spectrum.reflectance_percent,
                "metrics": metrics,
            }
        )
    return total_sse, evaluations


def information_criteria(sse: float, n_obs: int, parameter_count: int) -> dict[str, float]:
    if sse <= 0 or n_obs <= 0:
        return {"aic": float("nan"), "bic": float("nan")}
    likelihood_term = n_obs * np.log(sse / n_obs)
    return {
        "aic": float(likelihood_term + 2.0 * parameter_count),
        "bic": float(likelihood_term + np.log(n_obs) * parameter_count),
    }


def fit_joint(
    spectra: tuple[Spectrum, Spectrum],
    initialisation: dict[str, Any],
    use_calibration: bool = True,
    seeds: tuple[int, ...] = (2025, 2718, 3141),
    maxiter: int = 65,
) -> dict[str, Any]:
    """Run fixed-seed DE exploration and full-data L-BFGS-B refinement."""

    stride = 12
    subsampled = tuple(
        replace(
            item,
            sigma_cm_inv=item.sigma_cm_inv[::stride],
            reflectance_percent=item.reflectance_percent[::stride],
        )
        for item in spectra
    )
    n_sub = sum(item.reflectance_percent.size for item in subsampled)
    n_full = sum(item.reflectance_percent.size for item in spectra)

    def objective(data: tuple[Spectrum, Spectrum], unit: np.ndarray) -> float:
        physical = decode(unit, PARAMETER_BOUNDS)
        sse, _ = evaluate(physical, data, use_calibration=use_calibration)
        return float(sse / n_sub if data is subsampled else sse / n_full) if np.isfinite(sse) else 1e15

    rough_d = float(
        np.mean(
            [
                initialisation["attachment_1"]["d_initial_um"],
                initialisation["attachment_2"]["d_initial_um"],
            ]
        )
    )
    default_initial = np.array([rough_d, 15.0, 15.0, 600.0, 800.0, 1300.0, 800.0])
    candidates: list[tuple[str, np.ndarray, float]] = []
    seed_records: list[dict[str, Any]] = []
    for seed in seeds:
        de_result = differential_evolution(
            lambda unit: objective(subsampled, unit),
            bounds=[(0.0, 1.0)] * len(PARAMETER_BOUNDS),
            seed=seed,
            maxiter=maxiter,
            popsize=8,
            tol=5e-4,
            polish=False,
            workers=1,
            updating="immediate",
        )
        local_result = minimize(
            lambda unit: objective(subsampled, unit),
            de_result.x,
            method="L-BFGS-B",
            bounds=[(0.0, 1.0)] * len(PARAMETER_BOUNDS),
            options={"maxiter": 700, "ftol": 1e-12, "gtol": 1e-9, "maxls": 50},
        )
        local_unit = np.asarray(local_result.x, dtype=float)
        local_loss = float(objective(subsampled, local_unit))
        candidates.append((f"seed_{seed}", local_unit, local_loss))
        seed_records.append(
            {
                "seed": int(seed),
                "de_subsample_mse_percent_squared": float(de_result.fun),
                "local_subsample_mse_percent_squared": local_loss,
                "local_parameters": decode(local_unit, PARAMETER_BOUNDS),
                "local_success": bool(local_result.success),
                "local_message": str(local_result.message),
                "local_nit": int(getattr(local_result, "nit", 0)),
                "local_nfev": int(getattr(local_result, "nfev", 0)),
            }
        )

    initial_unit = encode(default_initial, PARAMETER_BOUNDS)
    candidates.append(("deterministic_period_initial", initial_unit, objective(subsampled, initial_unit)))
    start_name, start_unit, start_loss = min(candidates, key=lambda item: item[2])
    full_result = minimize(
        lambda unit: objective(spectra, unit),
        start_unit,
        method="L-BFGS-B",
        bounds=[(0.0, 1.0)] * len(PARAMETER_BOUNDS),
        options={"maxiter": 1600, "ftol": 1e-13, "gtol": 1e-10, "maxls": 70},
    )
    best_unit = np.asarray(full_result.x, dtype=float)
    parameters = decode(best_unit, PARAMETER_BOUNDS)
    sse, evaluations = evaluate(parameters, spectra, use_calibration=use_calibration)
    full_mse = float(sse / n_full)
    parameter_count = len(PARAMETER_BOUNDS) + (4 if use_calibration else 0)
    boundary_flags = {
        name: bool(
            abs(value - lower) / max(abs(upper - lower), 1.0) < 1e-4
            or abs(value - upper) / max(abs(upper - lower), 1.0) < 1e-4
        )
        for name, value, (lower, upper) in zip(PARAMETER_NAMES, parameters, PARAMETER_BOUNDS)
    }
    return {
        "model": "SiC_two_beam_Fresnel_Lorentz_Drude",
        "use_affine_calibration": bool(use_calibration),
        "parameters": dict(zip(PARAMETER_NAMES, parameters)),
        "calibration": {
            f"attachment_{row['attachment']}": {
                "scale": row["scale"],
                "offset_percent": row["offset_percent"],
            }
            for row in evaluations
        },
        "metrics": {
            f"attachment_{row['attachment']}": row["metrics"] for row in evaluations
        },
        "n_observations": int(n_full),
        "parameter_count_for_ic": int(parameter_count),
        "information_criteria": information_criteria(sse, n_full, parameter_count),
        "bounds": dict(zip(PARAMETER_NAMES, PARAMETER_BOUNDS)),
        "boundary_flags": boundary_flags,
        "optimizer": {
            "global_method": "differential_evolution",
            "local_method": "L-BFGS-B",
            "subsample_stride": int(stride),
            "seeds": list(seeds),
            "selected_subsample_start": start_name,
            "selected_subsample_mse_percent_squared": float(start_loss),
            "full_success": bool(full_result.success),
            "full_message": str(full_result.message),
            "full_nit": int(getattr(full_result, "nit", 0)),
            "full_nfev": int(getattr(full_result, "nfev", 0)),
            "full_mse_percent_squared": full_mse,
            "seed_records": seed_records,
        },
        "initialisation": initialisation,
        "_unit_parameters": best_unit,
        "_evaluations": evaluations,
    }


def profile_thickness(
    fit: dict[str, Any],
    spectra: tuple[Spectrum, Spectrum],
    points: int = 25,
    span_um: float = 0.45,
) -> list[dict[str, Any]]:
    """Re-optimise nuisance parameters along a fixed-thickness grid."""

    best_d = float(fit["parameters"]["d_um"])
    best_unit = np.asarray(fit["_unit_parameters"], dtype=float)
    lower = max(PARAMETER_BOUNDS[0][0], best_d - span_um)
    upper = min(PARAMETER_BOUNDS[0][1], best_d + span_um)
    n_full = sum(item.reflectance_percent.size for item in spectra)
    rows: list[dict[str, Any]] = []
    for d_value in np.linspace(lower, upper, points):
        d_unit = encode([d_value], (PARAMETER_BOUNDS[0],))[0]
        start_nuisance = best_unit[1:].copy()

        def objective_nuisance(nuisance: np.ndarray) -> float:
            unit = np.r_[d_unit, nuisance]
            parameters = decode(unit, PARAMETER_BOUNDS)
            sse, _ = evaluate(parameters, spectra, fit["use_affine_calibration"])
            return float(sse / n_full) if np.isfinite(sse) else 1e15

        result = minimize(
            objective_nuisance,
            start_nuisance,
            method="L-BFGS-B",
            bounds=[(0.0, 1.0)] * (len(PARAMETER_BOUNDS) - 1),
            options={"maxiter": 700, "ftol": 1e-12, "gtol": 1e-9, "maxls": 60},
        )
        rows.append(
            {
                "d_um": float(d_value),
                "mse_percent_squared": float(result.fun),
                "relative_to_best_mse": float(
                    result.fun / fit["optimizer"]["full_mse_percent_squared"] - 1.0
                ),
                "nuisance_optimisation_success": bool(result.success),
            }
        )
    rows.append(
        {
            "d_um": best_d,
            "mse_percent_squared": float(fit["optimizer"]["full_mse_percent_squared"]),
            "relative_to_best_mse": 0.0,
            "nuisance_optimisation_success": True,
        }
    )
    return sorted(rows, key=lambda row: row["d_um"])


def interpolate_crossing(left: dict[str, Any], right: dict[str, Any], threshold: float) -> float:
    x0, y0 = float(left["d_um"]), float(left["relative_to_best_mse"])
    x1, y1 = float(right["d_um"]), float(right["relative_to_best_mse"])
    if np.isclose(y1, y0):
        return float((x0 + x1) / 2.0)
    return float(x0 + (threshold - y0) * (x1 - x0) / (y1 - y0))


def profile_interval(rows: list[dict[str, Any]], threshold: float = 0.01) -> dict[str, Any]:
    ordered = sorted(rows, key=lambda row: row["d_um"])
    best_index = min(
        range(len(ordered)), key=lambda index: ordered[index]["relative_to_best_mse"]
    )
    left_cross = float(ordered[0]["d_um"])
    right_cross = float(ordered[-1]["d_um"])
    for index in range(best_index, 0, -1):
        if ordered[index - 1]["relative_to_best_mse"] >= threshold:
            left_cross = interpolate_crossing(ordered[index - 1], ordered[index], threshold)
            break
    for index in range(best_index, len(ordered) - 1):
        if ordered[index + 1]["relative_to_best_mse"] >= threshold:
            right_cross = interpolate_crossing(ordered[index], ordered[index + 1], threshold)
            break
    return {
        "relative_mse_threshold": float(threshold),
        "interval_um": [left_cross, right_cross],
        "half_width_um": float((right_cross - left_cross) / 2.0),
        "relative_half_width": float((right_cross - left_cross) / (2.0 * ordered[best_index]["d_um"])),
        "interpretation": "Re-optimised nuisance-parameter sensitivity interval; not a formal sampling confidence interval.",
    }


def sensitivity_constants(
    fit: dict[str, Any],
    spectra: tuple[Spectrum, Spectrum],
    relative_change: float = 0.02,
) -> list[dict[str, Any]]:
    """Conditionally perturb constants and re-optimise d on the local branch.

    A global one-dimensional search is inappropriate here: the interference
    phase is periodic and can contain neighbouring thickness branches.  The
    sensitivity check therefore stays within +/-0.45 um of the fitted value,
    the same local neighbourhood used by the thickness profile.
    """

    fixed_nuisance = np.asarray(list(fit["parameters"].values())[1:], dtype=float)
    n_full = sum(item.reflectance_percent.size for item in spectra)
    best_d = float(fit["parameters"]["d_um"])
    local_bounds = (
        max(PARAMETER_BOUNDS[0][0], best_d - 0.45),
        min(PARAMETER_BOUNDS[0][1], best_d + 0.45),
    )
    records: list[dict[str, Any]] = []
    for name in ("eps_inf", "omega_L_cm_inv", "omega_T_cm_inv"):
        for sign in (-1.0, 1.0):
            changed = dict(SIC_CONSTANTS)
            changed[name] *= 1.0 + sign * relative_change

            def conditional_loss(d_value: float) -> float:
                parameters = np.r_[d_value, fixed_nuisance]
                sse, _ = evaluate(
                    parameters,
                    spectra,
                    fit["use_affine_calibration"],
                    constants=changed,
                )
                return float(sse / n_full) if np.isfinite(sse) else 1e15

            result = minimize_scalar(
                conditional_loss,
                bounds=local_bounds,
                method="bounded",
                options={"xatol": 1e-5},
            )
            records.append(
                {
                    "perturbed_parameter": name,
                    "relative_change": float(sign * relative_change),
                    "changed_value": float(changed[name]),
                    "conditional_d_um": float(result.x),
                    "relative_d_change": float(result.x / fit["parameters"]["d_um"] - 1.0),
                    "conditional_mse_percent_squared": float(result.fun),
                    "local_search_bounds_um": list(local_bounds),
                }
            )
    return records


def round_trip_strength(
    spectra: tuple[Spectrum, Spectrum], parameters: np.ndarray
) -> dict[str, Any]:
    """Measure the amplitude multiplier of one additional internal round trip."""

    output: dict[str, Any] = {}
    for spectrum in spectra:
        sigma = spectrum.sigma_cm_inv
        gamma_l, gamma_t = parameters[1], parameters[2]
        n_layer = sic_refractive_index(
            sigma, gamma_l, gamma_t, parameters[3], parameters[4]
        )
        n_substrate = sic_refractive_index(
            sigma, gamma_l, gamma_t, parameters[5], parameters[6]
        )
        theta0 = np.deg2rad(spectrum.angle_deg)
        _, _, _, _, cos1 = fresnel_amplitudes(1.0, n_layer, theta0)
        theta1 = np.arccos(cos1)
        r12s, r12p, _, _, _ = fresnel_amplitudes(n_layer, n_substrate, theta1)
        r10s, r10p, _, _, _ = fresnel_amplitudes(n_layer, 1.0, theta1)
        delta = 4.0 * np.pi * np.asarray(sigma, dtype=np.complex128) * (
            parameters[0] * 1e-4
        ) * n_layer * cos1
        propagation = np.exp(1j * delta)
        rho = 0.5 * (
            np.abs(r10s * r12s * propagation)
            + np.abs(r10p * r12p * propagation)
        )
        output[f"attachment_{spectrum.attachment}"] = {
            "median": float(np.median(rho)),
            "p95": float(np.quantile(rho, 0.95)),
            "maximum": float(np.max(rho)),
        }
    return output


def write_predictions(
    fit: dict[str, Any], spectra: tuple[Spectrum, Spectrum], path: Path
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    by_attachment = {item.attachment: item for item in spectra}
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "attachment",
                "angle_deg",
                "wavenumber_cm_inv",
                "observed_percent",
                "physical_model_percent",
                "calibration_scale",
                "calibration_offset_percent",
                "fitted_percent",
                "residual_percent",
            ]
        )
        for row in fit["_evaluations"]:
            spectrum = by_attachment[row["attachment"]]
            for sigma, observed, physical, predicted in zip(
                spectrum.sigma_cm_inv,
                spectrum.reflectance_percent,
                row["physical_percent"],
                row["prediction_percent"],
            ):
                writer.writerow(
                    [
                        row["attachment"],
                        row["angle_deg"],
                        float(sigma),
                        float(observed),
                        float(physical),
                        float(row["scale"]),
                        float(row["offset_percent"]),
                        float(predicted),
                        float(predicted - observed),
                    ]
                )


def write_profile(rows: list[dict[str, Any]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def build_payload(
    spectra: tuple[Spectrum, Spectrum],
    audit: dict[str, Any],
    initialisation: dict[str, Any],
    fit: dict[str, Any],
    profile: list[dict[str, Any]],
    sensitivity: list[dict[str, Any]],
) -> dict[str, Any]:
    parameters = np.asarray(list(fit["parameters"].values()), dtype=float)
    public_fit = {key: value for key, value in fit.items() if not key.startswith("_")}
    return {
        "problem": "2025 CUMCM B Question 2",
        "input_audit": audit,
        "model_assumptions": {
            "geometry": "air--SiC epitaxial layer--SiC substrate",
            "interference": "Question-1 two-beam truncation",
            "polarisation": "unpolarised; arithmetic mean of s and p intensities",
            "shared_parameters_across_angles": [
                "d_um",
                "gamma_L_cm_inv",
                "gamma_T_cm_inv",
                "omega_p_layer_cm_inv",
                "gamma_p_layer_cm_inv",
                "omega_p_substrate_cm_inv",
                "gamma_p_substrate_cm_inv",
            ],
            "layer_substrate_distinction": "The two Drude pairs distinguish the epitaxial layer from the substrate; they are still shared across the two incident angles.",
            "affine_calibration": "per-angle bounded scale and offset eliminated by least squares",
            "constants": SIC_CONSTANTS,
        },
        "question_1_period_initialisation": initialisation,
        "joint_fit": public_fit,
        "thickness_profile": {
            "csv": "q2_thickness_profile.csv",
            "summary": profile_interval(profile, threshold=0.01),
        },
        "material_constant_sensitivity": {
            "relative_change": 0.02,
            "records": sensitivity,
        },
        "omitted_higher_return_diagnostic": round_trip_strength(spectra, parameters),
        "deliverables": {
            "prediction_csv": "q2_predictions.csv",
            "profile_csv": "q2_thickness_profile.csv",
            "fit_json": "q2_fit.json",
        },
        "environment": {
            "python": sys.version.split()[0],
            "platform": platform.platform(),
            "numpy": np.__version__,
        },
    }


def run(data_dir: Path = DEFAULT_DATA_DIR) -> dict[str, Any]:
    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    spectra = (load_spectrum(1, data_dir), load_spectrum(2, data_dir))
    audit = audit_spectra(spectra)
    initialisation = {
        f"attachment_{item.attachment}": period_initialisation(item)
        for item in spectra
    }
    write_json(RESULT_DIR / "q1_period_initialisation.json", initialisation)

    fit = fit_joint(spectra, initialisation, use_calibration=True)
    profile = profile_thickness(fit, spectra)
    write_profile(profile, RESULT_DIR / "q2_thickness_profile.csv")
    sensitivity = sensitivity_constants(fit, spectra)
    write_predictions(fit, spectra, RESULT_DIR / "q2_predictions.csv")

    payload = build_payload(spectra, audit, initialisation, fit, profile, sensitivity)
    write_json(RESULT_DIR / "q2_fit.json", payload)

    final_d = fit["parameters"]["d_um"]
    print(f"Q2 joint thickness: {final_d:.9f} um")
    print(
        "Q2 R2:",
        {
            key: value["r2"]
            for key, value in fit["metrics"].items()
        },
    )
    print("Q2 profile interval:", payload["thickness_profile"]["summary"]["interval_um"])
    return payload


def main() -> None:
    parser = argparse.ArgumentParser(description="Solve Question 2 for the 2025 B problem")
    parser.add_argument(
        "--data-dir",
        type=Path,
        default=DEFAULT_DATA_DIR,
        help="directory containing 附件1.xlsx and 附件2.xlsx",
    )
    args = parser.parse_args()
    run(args.data_dir)


if __name__ == "__main__":
    main()
