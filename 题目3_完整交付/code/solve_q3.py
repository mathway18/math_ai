"""Independent, reproducible solution for Question 3 of 2025 CUMCM B.

The program has two responsibilities required by Question 3:

1. Fit the silicon spectra (Attachments 3 and 4) with the two-beam model,
   the exact coherent Airy continuation, and a partially coherent geometric
   continuation.  The three models use the same data band and the same
   bounded measurement calibration, so their BIC comparison is auditable.
2. Re-evaluate the SiC spectra (Attachments 1 and 2) with the exact Airy
   continuation near the verified Question-2 solution.  This is a local
   correction check: it tests whether omitted higher-order returns can move
   the Question-2 thickness when the fitted material parameters are allowed
   to relax.

No workbook is modified.  Only the first endpoint whose measured reflectance
is exactly zero is removed from a numerical fitting array; the original raw
workbooks remain in data/.

Run from this directory or from any working directory:

    C:\\Python313\\python.exe code\\solve_q3.py

The code uses wavenumber in cm^-1, thickness in um, and the convention
exp(+i*delta).  Complex square roots are therefore selected to have a
non-negative attenuation component.
"""

from __future__ import annotations

import csv
import hashlib
import json
import platform
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Iterable

import numpy as np
from openpyxl import load_workbook
from scipy.optimize import differential_evolution, minimize
from scipy.signal import find_peaks, savgol_filter


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
PROCESSED_DIR = DATA_DIR / "processed"
RESULTS_DIR = ROOT / "results"

ANGLES = {1: 10.0, 2: 15.0, 3: 10.0, 4: 15.0}
MATERIALS = {1: "SiC", 2: "SiC", 3: "Si", 4: "Si"}

# The common silicon fit band is deliberately inside the documented validity
# range of the Sellmeier background and avoids the low-wavenumber extrapolation
# of that empirical formula.  The final narrower band is selected by a
# dual-angle fringe-stability diagnostic below.
SI_VALID_BAND_CM_INV = (1000.0, 3600.0)
SI_FIT_BAND_CANDIDATES = tuple(
    (low, low + 1400.0)
    for low in (1000.0, 1200.0, 1400.0, 1600.0, 1800.0, 2000.0)
)

# Silicon Sellmeier coefficients, with wavelength in micrometres.  They are
# recorded as a modelling assumption because the contest attachments provide
# reflectance only, not a complex refractive-index table.
SI_SELLMEIER_B = (10.6684293, 0.0030434748, 1.54133408)
SI_SELLMEIER_C = (0.0909121907, 1.287660172, 1218816.0)
SI_SELLMEIER_WAVELENGTH_UM = (1.357, 11.04)

# The following constants are the same 4H-SiC Lorentz--Drude assumptions used
# in the independently delivered Question-2 package.  Q3 uses them only for
# the requested higher-order-return correction check.
SIC_CONSTANTS = {
    "eps_inf": 6.56,
    "omega_L_cm_inv": 970.0,
    "omega_T_cm_inv": 798.0,
}

SIC_PARAMETER_NAMES = (
    "d_um",
    "gamma_L_cm_inv",
    "gamma_T_cm_inv",
    "omega_p_layer_cm_inv",
    "gamma_p_layer_cm_inv",
    "omega_p_substrate_cm_inv",
    "gamma_p_substrate_cm_inv",
)
SIC_BOUNDS = (
    (3.0, 12.0),
    (0.1, 1500.0),
    (0.1, 1500.0),
    (1.0, 5000.0),
    (0.1, 2000.0),
    (1.0, 5000.0),
    (0.1, 2000.0),
)

# Fallback reference in case the sibling Question-2 package is moved.  When
# it exists, the program reads the verified JSON result instead and records the
# source path in q3_result.json.
SIC_Q2_FALLBACK = np.array(
    [
        7.400452153256163,
        0.1,
        2.5580835553030545,
        473.8485626664637,
        1184.1327484848778,
        1123.2584342176108,
        611.5947931686339,
    ],
    dtype=float,
)


@dataclass(frozen=True)
class Spectrum:
    attachment: int
    angle_deg: float
    material: str
    sigma_cm_inv: np.ndarray
    reflectance_percent: np.ndarray
    source_rows: int
    removed_initial_zero: bool


def json_ready(value: Any) -> Any:
    """Convert NumPy values and non-finite floats to JSON-safe values."""

    if isinstance(value, dict):
        return {str(key): json_ready(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [json_ready(item) for item in value]
    if isinstance(value, np.ndarray):
        return [json_ready(item) for item in value.tolist()]
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, np.floating):
        return float(value) if np.isfinite(value) else None
    if isinstance(value, float):
        return value if np.isfinite(value) else None
    if isinstance(value, np.bool_):
        return bool(value)
    return value


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(json_ready(payload), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def read_workbook(attachment: int) -> tuple[np.ndarray, list[str]]:
    """Read the first sheet's first two columns as floating-point data."""

    path = DATA_DIR / f"附件{attachment}.xlsx"
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(item) for item in header_row[:2]]
    rows = [
        (row[0], row[1])
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row[0] is not None and row[1] is not None
    ]
    workbook.close()
    return np.asarray(rows, dtype=float), headers


def load_spectrum(attachment: int) -> Spectrum:
    """Load one attachment and remove only its anomalous first zero endpoint."""

    values, _ = read_workbook(attachment)
    sigma = values[:, 0]
    reflectance = values[:, 1]
    removed = bool(reflectance.size and np.isclose(reflectance[0], 0.0, atol=1e-12))
    if removed:
        sigma = sigma[1:]
        reflectance = reflectance[1:]
    return Spectrum(
        attachment=attachment,
        angle_deg=ANGLES[attachment],
        material=MATERIALS[attachment],
        sigma_cm_inv=sigma,
        reflectance_percent=reflectance,
        source_rows=int(values.shape[0]),
        removed_initial_zero=removed,
    )


def audit_inputs(spectra: dict[int, Spectrum]) -> dict[str, Any]:
    """Audit all four workbooks and export immutable cleaned CSV copies."""

    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    entries: list[dict[str, Any]] = []
    for attachment in sorted(spectra):
        raw, headers = read_workbook(attachment)
        sigma_raw, y_raw = raw[:, 0], raw[:, 1]
        spacing = np.diff(sigma_raw)
        spectrum = spectra[attachment]
        cleaned_path = PROCESSED_DIR / f"附件{attachment}_clean.csv"
        with cleaned_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(["wavenumber_cm_inv", "reflectance_percent"])
            writer.writerows(
                zip(spectrum.sigma_cm_inv.tolist(), spectrum.reflectance_percent.tolist())
            )
        entries.append(
            {
                "attachment": attachment,
                "material": spectrum.material,
                "angle_deg": spectrum.angle_deg,
                "source_file": f"data/附件{attachment}.xlsx",
                "source_sha256": sha256(DATA_DIR / f"附件{attachment}.xlsx"),
                "headers": headers,
                "source_rows": int(raw.shape[0]),
                "fitting_rows": int(spectrum.sigma_cm_inv.size),
                "removed_first_zero": bool(spectrum.removed_initial_zero),
                "wavenumber_cm_inv": {
                    "min": float(sigma_raw.min()),
                    "max": float(sigma_raw.max()),
                    "median_step": float(np.median(spacing)),
                    "min_step": float(spacing.min()),
                    "max_step": float(spacing.max()),
                    "duplicate_count": int(np.sum(np.isclose(spacing, 0.0))),
                },
                "reflectance_percent": {
                    "min": float(y_raw.min()),
                    "max": float(y_raw.max()),
                    "mean": float(y_raw.mean()),
                    "std": float(y_raw.std()),
                    "zero_count": int(np.sum(np.isclose(y_raw, 0.0))),
                    "above_100_count": int(np.sum(y_raw > 100.0)),
                },
                "processed_csv": str(cleaned_path.relative_to(ROOT)),
            }
        )
    payload = {
        "rule": "Raw workbooks are unchanged; only an exact first 0% endpoint is removed in memory for fitting.",
        "attachments": entries,
    }
    write_json(RESULTS_DIR / "data_audit.json", payload)
    return payload


def _forward_root(value: np.ndarray | complex) -> np.ndarray:
    """Select a refractive-index branch with non-negative attenuation."""

    root = np.sqrt(np.asarray(value, dtype=np.complex128))
    flip_real = (np.real(root) < 0) | (
        (np.abs(np.real(root)) < 1e-14) & (np.imag(root) < 0)
    )
    root = np.where(flip_real, -root, root)
    # exp(+i delta) attenuates when Im(n) >= 0.
    root = np.where(np.imag(root) < 0, np.conj(root), root)
    return root


def _forward_cosine(value: np.ndarray | complex) -> np.ndarray:
    cosine = np.sqrt(np.asarray(value, dtype=np.complex128))
    flip = (np.real(cosine) < 0) | (
        (np.abs(np.real(cosine)) < 1e-14) & (np.imag(cosine) < 0)
    )
    return np.where(flip, -cosine, cosine)


def fresnel_amplitudes(
    n_i: np.ndarray | complex,
    n_j: np.ndarray | complex,
    theta_i_rad: float | np.ndarray,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    """Return s/p reflection and transmission amplitudes and cos(theta_j)."""

    n_i = np.asarray(n_i, dtype=np.complex128)
    n_j = np.asarray(n_j, dtype=np.complex128)
    theta_i = np.asarray(theta_i_rad, dtype=np.complex128)
    cos_i = np.cos(theta_i)
    sin_j = n_i * np.sin(theta_i) / n_j
    cos_j = _forward_cosine(1.0 - sin_j**2)

    den_s = n_i * cos_i + n_j * cos_j
    den_p = n_j * cos_i + n_i * cos_j
    r_s = (n_i * cos_i - n_j * cos_j) / den_s
    t_s = 2.0 * n_i * cos_i / den_s
    r_p = (n_j * cos_i - n_i * cos_j) / den_p
    t_p = 2.0 * n_i * cos_i / den_p
    return r_s, r_p, t_s, t_p, cos_j


def layered_reflectance_percent(
    sigma_cm_inv: np.ndarray,
    n_layer: np.ndarray,
    n_substrate: np.ndarray | complex,
    d_um: float,
    angle_deg: float,
    eta: float,
) -> np.ndarray:
    """Air--layer--substrate reflectance with a controlled return series.

    Let ``delta = 4*pi*sigma*n1*cos(theta1)*d`` be the round-trip phase.
    The first-return (Question-1) truncation is

        A = r01 + t01*r12*t10*exp(i*delta).

    Repeated internal returns form a geometric series with ratio
    ``r10*r12*exp(i*delta)``.  ``eta=1`` is the exact coherent Airy sum;
    ``eta=0`` is the first-return truncation; an intermediate eta is a
    transparent partial-coherence/loss parameter used only for model
    diagnosis.
    """

    if not 0.0 <= float(eta) <= 1.0:
        raise ValueError("eta must be between 0 and 1")
    sigma = np.asarray(sigma_cm_inv, dtype=np.complex128)
    n1 = np.asarray(n_layer, dtype=np.complex128)
    n2 = np.asarray(n_substrate, dtype=np.complex128)
    theta0 = np.deg2rad(float(angle_deg))

    r01s, r01p, t01s, t01p, cos1 = fresnel_amplitudes(1.0, n1, theta0)
    theta1 = np.arccos(cos1)
    r12s, r12p, _, _, _ = fresnel_amplitudes(n1, n2, theta1)
    r10s, r10p, t10s, t10p, _ = fresnel_amplitudes(n1, 1.0, theta1)

    delta = 4.0 * np.pi * sigma * (float(d_um) * 1e-4) * n1 * cos1
    propagation = np.exp(1j * delta)

    def total_amplitude(
        r01: np.ndarray,
        r12: np.ndarray,
        r10: np.ndarray,
        t01: np.ndarray,
        t10: np.ndarray,
    ) -> np.ndarray:
        first_return = t01 * r12 * t10 * propagation
        if float(eta) == 0.0:
            return r01 + first_return
        denominator = 1.0 - float(eta) * r10 * r12 * propagation
        return r01 + first_return / denominator

    amplitude_s = total_amplitude(r01s, r12s, r10s, t01s, t10s)
    amplitude_p = total_amplitude(r01p, r12p, r10p, t01p, t10p)
    result = 50.0 * (np.abs(amplitude_s) ** 2 + np.abs(amplitude_p) ** 2)
    return np.real(result)


def si_sellmeier_index(sigma_cm_inv: np.ndarray) -> np.ndarray:
    sigma = np.asarray(sigma_cm_inv, dtype=float)
    wavelength_um = 1e4 / sigma
    low, high = SI_SELLMEIER_WAVELENGTH_UM
    if np.any((wavelength_um < low - 1e-12) | (wavelength_um > high + 1e-12)):
        raise ValueError(
            f"Si Sellmeier relation is valid only for {low:g}--{high:g} um"
        )
    wavelength_sq = wavelength_um**2
    n_sq = 1.0 + sum(
        b * wavelength_sq / (wavelength_sq - c)
        for b, c in zip(SI_SELLMEIER_B, SI_SELLMEIER_C)
    )
    return np.sqrt(n_sq)


def si_drude_index(
    sigma_cm_inv: np.ndarray,
    omega_p_cm_inv: float,
    gamma_p_cm_inv: float = 0.0,
) -> np.ndarray:
    """Silicon background plus a free-carrier Drude correction."""

    sigma = np.asarray(sigma_cm_inv, dtype=np.complex128)
    background_eps = si_sellmeier_index(np.real(sigma)) ** 2
    drude = omega_p_cm_inv**2 / (
        sigma**2 + 1j * float(gamma_p_cm_inv) * sigma
    )
    return _forward_root(background_eps - drude)


def sic_dielectric(
    sigma_cm_inv: np.ndarray,
    gamma_l: float,
    gamma_t: float,
    omega_p: float,
    gamma_p: float,
) -> np.ndarray:
    sigma = np.asarray(sigma_cm_inv, dtype=np.complex128)
    c = SIC_CONSTANTS
    lattice = c["eps_inf"] * (
        (c["omega_L_cm_inv"] ** 2 - sigma**2 - 1j * gamma_l * sigma)
        / (c["omega_T_cm_inv"] ** 2 - sigma**2 - 1j * gamma_t * sigma)
    )
    drude = omega_p**2 / (sigma**2 + 1j * gamma_p * sigma)
    return lattice - drude


def sic_index(
    sigma_cm_inv: np.ndarray,
    gamma_l: float,
    gamma_t: float,
    omega_p: float,
    gamma_p: float,
) -> np.ndarray:
    return _forward_root(sic_dielectric(sigma_cm_inv, gamma_l, gamma_t, omega_p, gamma_p))


def affine_calibration(
    physical_percent: np.ndarray,
    observed_percent: np.ndarray,
    scale_bounds: tuple[float, float] = (0.25, 1.75),
    offset_bounds: tuple[float, float] = (-30.0, 30.0),
) -> tuple[float, float, np.ndarray]:
    """Solve the two linear measurement nuisance parameters analytically."""

    physical = np.asarray(physical_percent, dtype=float)
    observed = np.asarray(observed_percent, dtype=float)
    design = np.column_stack((physical, np.ones_like(physical)))
    scale, offset = np.linalg.lstsq(design, observed, rcond=None)[0]
    scale = float(np.clip(scale, *scale_bounds))
    offset = float(np.clip(np.mean(observed - scale * physical), *offset_bounds))
    return scale, offset, scale * physical + offset


def fit_metrics(observed: np.ndarray, predicted: np.ndarray) -> dict[str, float]:
    residual = np.asarray(predicted, dtype=float) - np.asarray(observed, dtype=float)
    sse = float(np.sum(residual**2))
    sst = float(np.sum((observed - np.mean(observed)) ** 2))
    absolute = np.abs(residual)
    return {
        "sse_percent_squared": sse,
        "rmse_percent": float(np.sqrt(np.mean(residual**2))),
        "mae_percent": float(np.mean(absolute)),
        "max_abs_residual_percent": float(np.max(absolute)),
        "mean_residual_percent": float(np.mean(residual)),
        "residual_q50_abs_percent": float(np.quantile(absolute, 0.50)),
        "residual_q95_abs_percent": float(np.quantile(absolute, 0.95)),
        "r2": float(1.0 - sse / sst) if sst > 0 else float("nan"),
    }


def information_criteria(sse: float, n_obs: int, parameter_count: int) -> dict[str, float]:
    log_term = n_obs * np.log(sse / n_obs)
    return {
        "aic": float(log_term + 2.0 * parameter_count),
        "bic": float(log_term + np.log(n_obs) * parameter_count),
    }


def round_trip_strength(
    sigma_cm_inv: np.ndarray,
    n_layer: np.ndarray,
    n_substrate: np.ndarray,
    d_um: float,
    angle_deg: float,
) -> np.ndarray:
    """Magnitude of the extra-return ratio |r10*r12*exp(i delta)|."""

    sigma = np.asarray(sigma_cm_inv, dtype=np.complex128)
    n1 = np.asarray(n_layer, dtype=np.complex128)
    n2 = np.asarray(n_substrate, dtype=np.complex128)
    theta0 = np.deg2rad(float(angle_deg))
    _, _, _, _, cos1 = fresnel_amplitudes(1.0, n1, theta0)
    theta1 = np.arccos(cos1)
    r12s, r12p, _, _, _ = fresnel_amplitudes(n1, n2, theta1)
    r10s, r10p, _, _, _ = fresnel_amplitudes(n1, 1.0, theta1)
    delta = 4.0 * np.pi * sigma * (float(d_um) * 1e-4) * n1 * cos1
    propagation = np.exp(1j * delta)
    rho_s = np.abs(r10s * r12s * propagation)
    rho_p = np.abs(r10p * r12p * propagation)
    return 0.5 * (rho_s + rho_p)


def _odd_window(requested: int, available: int, minimum: int = 5) -> int:
    window = min(int(requested), int(available) - (1 - int(available) % 2))
    if window % 2 == 0:
        window -= 1
    if window < minimum:
        raise ValueError("not enough points for smoothing")
    return window


def _subset(spectrum: Spectrum, band: tuple[float, float]) -> Spectrum:
    mask = (spectrum.sigma_cm_inv >= band[0]) & (spectrum.sigma_cm_inv <= band[1])
    return Spectrum(
        attachment=spectrum.attachment,
        angle_deg=spectrum.angle_deg,
        material=spectrum.material,
        sigma_cm_inv=spectrum.sigma_cm_inv[mask],
        reflectance_percent=spectrum.reflectance_percent[mask],
        source_rows=spectrum.source_rows,
        removed_initial_zero=spectrum.removed_initial_zero,
    )


def _fft_estimate(spectrum: Spectrum, low: float, high: float) -> dict[str, float]:
    """Estimate thickness after remapping to the dispersion-corrected phase axis."""

    data = _subset(spectrum, (low, high))
    sigma = data.sigma_cm_inv
    response = data.reflectance_percent
    smooth = savgol_filter(response, _odd_window(801, response.size), 3)
    residual = response - smooth
    n = si_sellmeier_index(sigma)
    phase_axis = 2.0 * sigma * np.sqrt(n**2 - np.sin(np.deg2rad(data.angle_deg)) ** 2)
    uniform = np.linspace(float(phase_axis.min()), float(phase_axis.max()), 8192)
    signal = np.interp(uniform, phase_axis, residual)
    signal = (signal - np.mean(signal)) * np.hanning(signal.size)
    amplitude = np.abs(np.fft.rfft(signal))
    frequency = np.fft.rfftfreq(signal.size, d=uniform[1] - uniform[0])
    candidate = (frequency * 1e4 >= 1.0) & (frequency * 1e4 <= 8.0)
    indices = np.flatnonzero(candidate)
    peak_index = int(indices[np.argmax(amplitude[candidate])])
    noise = float(np.median(amplitude[candidate]))
    return {
        "band_cm_inv": [float(low), float(high)],
        "d_fft_um": float(frequency[peak_index] * 1e4),
        "peak_to_median": float(amplitude[peak_index] / noise),
    }


def select_si_band(spectra: tuple[Spectrum, Spectrum]) -> dict[str, Any]:
    """Choose a common band with a strong and angle-stable fringe frequency."""

    candidates: list[dict[str, Any]] = []
    for low, high in SI_FIT_BAND_CANDIDATES:
        first = _fft_estimate(spectra[0], low, high)
        second = _fft_estimate(spectra[1], low, high)
        gap = abs(first["d_fft_um"] - second["d_fft_um"]) / np.mean(
            [first["d_fft_um"], second["d_fft_um"]]
        )
        score = min(first["peak_to_median"], second["peak_to_median"]) / (1.0 + 20.0 * gap)
        candidates.append(
            {
                "band_cm_inv": [float(low), float(high)],
                "attachment_3": first,
                "attachment_4": second,
                "relative_angle_gap": float(gap),
                "score": float(score),
            }
        )
    selected = max(candidates, key=lambda item: item["score"])
    return {
        "sellmeier_valid_band_cm_inv": list(SI_VALID_BAND_CM_INV),
        "selection_rule": "Maximise the lower dual-angle FFT peak-to-median ratio after penalising angle disagreement.",
        "candidates": candidates,
        "selected": selected,
    }


def peak_initialisation(spectra: tuple[Spectrum, Spectrum], band: tuple[float, float]) -> dict[str, Any]:
    """Use adjacent peaks only to obtain a transparent numerical scale."""

    records: dict[str, Any] = {}
    estimates: list[float] = []
    for spectrum in spectra:
        data = _subset(spectrum, band)
        smooth = savgol_filter(data.reflectance_percent, _odd_window(101, data.reflectance_percent.size), 3)
        peaks, props = find_peaks(smooth, prominence=0.5, distance=250)
        locations = data.sigma_cm_inv[peaks]
        spacing = np.diff(locations)
        n0 = float(np.median(si_sellmeier_index(locations))) if locations.size else 3.42
        factor = np.sqrt(n0**2 - np.sin(np.deg2rad(spectrum.angle_deg)) ** 2)
        d_um = float(1e4 / (2.0 * factor * np.median(spacing))) if spacing.size else float("nan")
        estimates.append(d_um)
        records[f"attachment_{spectrum.attachment}"] = {
            "attachment": spectrum.attachment,
            "angle_deg": spectrum.angle_deg,
            "band_cm_inv": [float(band[0]), float(band[1])],
            "smoothing": {"method": "Savitzky-Golay", "window_points": 101, "polynomial_order": 3},
            "peaks_cm_inv": locations,
            "peak_prominence_percent": props.get("prominences", []),
            "successive_spacing_cm_inv": spacing,
            "median_spacing_cm_inv": float(np.median(spacing)) if spacing.size else float("nan"),
            "n_initial": n0,
            "d_initial_um": d_um,
        }
    return {
        "records": records,
        "mean_d_initial_um": float(np.nanmean(estimates)),
        "role": "Initial scale and diagnostic only; final thickness is obtained by spectral inversion.",
    }


def _decode(unit: np.ndarray, bounds: tuple[tuple[float, float], ...]) -> np.ndarray:
    lower = np.asarray([item[0] for item in bounds], dtype=float)
    upper = np.asarray([item[1] for item in bounds], dtype=float)
    return lower + np.clip(np.asarray(unit, dtype=float), 0.0, 1.0) * (upper - lower)


def _encode(physical: Iterable[float], bounds: tuple[tuple[float, float], ...]) -> np.ndarray:
    values = np.asarray(tuple(physical), dtype=float)
    lower = np.asarray([item[0] for item in bounds], dtype=float)
    upper = np.asarray([item[1] for item in bounds], dtype=float)
    return np.clip((values - lower) / (upper - lower), 0.0, 1.0)


def _optimise(
    objective_subsample: Callable[[np.ndarray], float],
    objective_full: Callable[[np.ndarray], float],
    bounds: tuple[tuple[float, float], ...],
    initial_physical: Iterable[float],
    seeds: tuple[int, ...] = (2025, 2718, 3141),
    maxiter: int = 70,
) -> tuple[np.ndarray, dict[str, Any]]:
    """Deterministic DE exploration followed by full-data L-BFGS-B."""

    unit_bounds = [(0.0, 1.0)] * len(bounds)
    records: list[dict[str, Any]] = []
    candidates: list[tuple[str, np.ndarray, float]] = []
    for seed in seeds:
        de = differential_evolution(
            objective_subsample,
            bounds=unit_bounds,
            seed=seed,
            maxiter=maxiter,
            popsize=8,
            tol=5e-4,
            polish=False,
            workers=1,
            updating="immediate",
        )
        local = minimize(
            objective_subsample,
            de.x,
            method="L-BFGS-B",
            bounds=unit_bounds,
            options={"maxiter": 900, "ftol": 1e-12, "gtol": 1e-9, "maxls": 60},
        )
        candidate = np.asarray(local.x, dtype=float)
        local_loss = float(objective_subsample(candidate))
        candidates.append((f"seed_{seed}", candidate, local_loss))
        records.append(
            {
                "seed": int(seed),
                "de_subsample_mse_percent_squared": float(de.fun),
                "local_subsample_mse_percent_squared": local_loss,
                "local_success": bool(local.success),
                "local_message": str(local.message),
            }
        )

    deterministic = _encode(initial_physical, bounds)
    candidates.append(("deterministic_initial", deterministic, float(objective_subsample(deterministic))))
    selected_name, selected_unit, selected_loss = min(candidates, key=lambda item: item[2])
    full = minimize(
        objective_full,
        selected_unit,
        method="L-BFGS-B",
        bounds=unit_bounds,
        options={"maxiter": 1800, "ftol": 1e-13, "gtol": 1e-10, "maxls": 70},
    )
    return np.asarray(full.x, dtype=float), {
        "global_method": "differential_evolution",
        "local_method": "L-BFGS-B",
        "seeds": list(seeds),
        "selected_subsample_start": selected_name,
        "selected_subsample_mse_percent_squared": float(selected_loss),
        "full_success": bool(full.success),
        "full_message": str(full.message),
        "full_nit": int(getattr(full, "nit", -1)),
        "full_nfev": int(getattr(full, "nfev", -1)),
        "full_mse_percent_squared": float(full.fun),
        "seed_records": records,
    }


def _si_predict(
    sigma: np.ndarray,
    angle_deg: float,
    physical: np.ndarray,
    eta: float,
) -> np.ndarray:
    d_um, wp_layer, wp_substrate = np.asarray(physical, dtype=float)[:3]
    n_layer = si_drude_index(sigma, wp_layer, 0.0)
    n_substrate = si_drude_index(sigma, wp_substrate, 0.0)
    return layered_reflectance_percent(sigma, n_layer, n_substrate, d_um, angle_deg, eta)


def _evaluate_si(
    physical: np.ndarray,
    eta: float,
    spectra: tuple[Spectrum, ...],
) -> tuple[float, list[dict[str, Any]]]:
    total_sse = 0.0
    evaluations: list[dict[str, Any]] = []
    for spectrum in spectra:
        physical_percent = _si_predict(
            spectrum.sigma_cm_inv, spectrum.angle_deg, physical, eta
        )
        if not np.all(np.isfinite(physical_percent)):
            return float("inf"), []
        scale, offset, prediction = affine_calibration(
            physical_percent, spectrum.reflectance_percent
        )
        metrics = fit_metrics(spectrum.reflectance_percent, prediction)
        total_sse += metrics["sse_percent_squared"]
        evaluations.append(
            {
                "attachment": spectrum.attachment,
                "angle_deg": spectrum.angle_deg,
                "scale": scale,
                "offset_percent": offset,
                "metrics": metrics,
                "physical_percent": physical_percent,
                "prediction_percent": prediction,
            }
        )
    return total_sse, evaluations


def _si_mode_spec(mode: str) -> tuple[tuple[tuple[float, float], ...], float | None]:
    base_bounds = ((2.0, 5.5), (0.0, 5000.0), (0.0, 5000.0))
    if mode == "two_beam":
        return base_bounds, 0.0
    if mode == "full_multibeam":
        return base_bounds, 1.0
    if mode == "partial_multibeam":
        return base_bounds + ((0.0, 1.0),), None
    raise ValueError(f"unknown silicon mode: {mode}")


def fit_si_model(
    mode: str,
    spectra: tuple[Spectrum, Spectrum],
    initial_d_um: float,
) -> dict[str, Any]:
    """Fit one Si model using exactly the same observations and calibration."""

    bounds, fixed_eta = _si_mode_spec(mode)
    dimension = len(bounds)
    stride = 6
    subsampled = tuple(
        Spectrum(
            attachment=item.attachment,
            angle_deg=item.angle_deg,
            material=item.material,
            sigma_cm_inv=item.sigma_cm_inv[::stride],
            reflectance_percent=item.reflectance_percent[::stride],
            source_rows=item.source_rows,
            removed_initial_zero=item.removed_initial_zero,
        )
        for item in spectra
    )

    def unpack(unit: np.ndarray) -> tuple[np.ndarray, float]:
        physical = _decode(unit, bounds)
        if fixed_eta is None:
            return physical[:3], float(physical[3])
        return physical[:3], float(fixed_eta)

    def objective(data: tuple[Spectrum, ...], unit: np.ndarray) -> float:
        physical, eta = unpack(unit)
        sse, _ = _evaluate_si(physical, eta, data)
        n = sum(item.reflectance_percent.size for item in data)
        return sse / n

    initial = [float(initial_d_um), 100.0, 4000.0]
    if fixed_eta is None:
        initial.append(0.70)
    unit, optimizer = _optimise(
        lambda candidate: objective(subsampled, candidate),
        lambda candidate: objective(spectra, candidate),
        bounds,
        initial,
        maxiter=70 if fixed_eta is None else 55,
    )
    physical, eta = unpack(unit)
    sse, evaluations = _evaluate_si(physical, eta, spectra)
    n_obs = sum(item.reflectance_percent.size for item in spectra)
    parameter_count = len(bounds) + 4  # physical parameters plus two scales/offsets

    n_by_attachment: dict[str, dict[str, float]] = {}
    for spectrum in spectra:
        n_layer = si_drude_index(spectrum.sigma_cm_inv, physical[1], 0.0)
        n_substrate = si_drude_index(spectrum.sigma_cm_inv, physical[2], 0.0)
        rho = round_trip_strength(
            spectrum.sigma_cm_inv,
            n_layer,
            n_substrate,
            physical[0],
            spectrum.angle_deg,
        )
        n_by_attachment[f"attachment_{spectrum.attachment}"] = {
            "median": float(np.median(rho)),
            "p95": float(np.quantile(rho, 0.95)),
            "maximum": float(np.max(rho)),
        }

    return {
        "mode": mode,
        "model_definition": {
            "eta": "0 for Question-1 two-beam truncation; 1 for exact coherent Airy sum; fitted in [0,1] for partial_multibeam.",
            "physical_parameters": ["d_um", "omega_p_layer_cm_inv", "omega_p_substrate_cm_inv"]
            + (["eta"] if fixed_eta is None else []),
            "calibration": "per-angle bounded scale and offset solved analytically inside every objective evaluation",
        },
        "parameters": {
            "d_um": float(physical[0]),
            "omega_p_layer_cm_inv": float(physical[1]),
            "omega_p_substrate_cm_inv": float(physical[2]),
            "eta": float(eta),
        },
        "calibration": {
            f"attachment_{item['attachment']}": {
                "scale": float(item["scale"]),
                "offset_percent": float(item["offset_percent"]),
            }
            for item in evaluations
        },
        "metrics": {
            f"attachment_{item['attachment']}": item["metrics"]
            for item in evaluations
        },
        "n_observations": int(n_obs),
        "parameter_count_for_ic": int(parameter_count),
        "information_criteria": information_criteria(sse, n_obs, parameter_count),
        "bounds": {
            name: list(bound)
            for name, bound in zip(
                ["d_um", "omega_p_layer_cm_inv", "omega_p_substrate_cm_inv"]
                + (["eta"] if fixed_eta is None else []),
                bounds,
            )
        },
        "optimizer": optimizer,
        "round_trip_strength": n_by_attachment,
        "_evaluations": evaluations,
    }


def individual_si_fit(
    spectrum: Spectrum,
    selected_model: str,
    initial_parameters: dict[str, float],
) -> dict[str, Any]:
    """Fit each angle separately as a consistency diagnostic."""

    bounds, fixed_eta = _si_mode_spec(selected_model)
    if fixed_eta is None:
        initial = [
            initial_parameters["d_um"],
            initial_parameters["omega_p_layer_cm_inv"],
            initial_parameters["omega_p_substrate_cm_inv"],
            initial_parameters["eta"],
        ]
    else:
        initial = [
            initial_parameters["d_um"],
            initial_parameters["omega_p_layer_cm_inv"],
            initial_parameters["omega_p_substrate_cm_inv"],
        ]

    def objective(unit: np.ndarray) -> float:
        physical = _decode(unit, bounds)
        eta = float(physical[3]) if fixed_eta is None else float(fixed_eta)
        sse, _ = _evaluate_si(physical[:3], eta, (spectrum,))
        return sse / spectrum.reflectance_percent.size

    result = minimize(
        objective,
        _encode(initial, bounds),
        method="L-BFGS-B",
        bounds=[(0.0, 1.0)] * len(bounds),
        options={"maxiter": 1200, "ftol": 1e-13, "gtol": 1e-10, "maxls": 70},
    )
    physical = _decode(result.x, bounds)
    eta = float(physical[3]) if fixed_eta is None else float(fixed_eta)
    return {
        "attachment": spectrum.attachment,
        "angle_deg": spectrum.angle_deg,
        "d_um": float(physical[0]),
        "omega_p_layer_cm_inv": float(physical[1]),
        "omega_p_substrate_cm_inv": float(physical[2]),
        "eta": eta,
        "mse_percent_squared": float(result.fun),
        "success": bool(result.success),
        "message": str(result.message),
    }


def si_substrate_sensitivity(
    spectra: tuple[Spectrum, Spectrum],
    selected: dict[str, Any],
) -> dict[str, Any]:
    """Conditionally re-fit thickness while fixing the substrate surrogate.

    The selected silicon fit reaches the upper bound of its substrate Drude
    surrogate.  This diagnostic deliberately tests values beyond that bound,
    because the purpose is to show how much the thickness depends on the
    unsupported surrogate, not to claim that those values are measured.
    """

    mode = str(selected["mode"])
    selected_parameters = selected["parameters"]
    fixed_eta = 0.0 if mode == "two_beam" else 1.0 if mode == "full_multibeam" else None
    reduced_bounds: tuple[tuple[float, float], ...] = (
        (2.0, 5.5),
        (0.0, 5000.0),
    )
    if fixed_eta is None:
        reduced_bounds += ((0.0, 1.0),)
    fixed_values = (3000.0, 4000.0, 5000.0, 6500.0, 8000.0)
    reference_mse = float(selected["optimizer"]["full_mse_percent_squared"])
    starts = [
        [
            selected_parameters["d_um"],
            selected_parameters["omega_p_layer_cm_inv"],
        ]
        + ([] if fixed_eta is not None else [selected_parameters["eta"]]),
        [3.30, 0.0] + ([] if fixed_eta is not None else [0.60]),
        [3.50, 1000.0] + ([] if fixed_eta is not None else [0.80]),
        [3.80, 3000.0] + ([] if fixed_eta is not None else [0.50]),
    ]

    records: list[dict[str, Any]] = []
    for fixed_substrate in fixed_values:
        def objective(unit: np.ndarray) -> float:
            reduced = _decode(unit, reduced_bounds)
            physical = np.array([reduced[0], reduced[1], fixed_substrate], dtype=float)
            eta = float(reduced[2]) if fixed_eta is None else float(fixed_eta)
            sse, _ = _evaluate_si(physical, eta, spectra)
            return sse / sum(item.reflectance_percent.size for item in spectra)

        candidates: list[tuple[np.ndarray, Any]] = []
        for start in starts:
            result = minimize(
                objective,
                _encode(start, reduced_bounds),
                method="L-BFGS-B",
                bounds=[(0.0, 1.0)] * len(reduced_bounds),
                options={"maxiter": 1200, "ftol": 1e-13, "gtol": 1e-10, "maxls": 70},
            )
            candidates.append((np.asarray(result.x, dtype=float), result))
        unit, result = min(candidates, key=lambda item: float(item[1].fun))
        reduced = _decode(unit, reduced_bounds)
        physical = np.array([reduced[0], reduced[1], fixed_substrate], dtype=float)
        eta = float(reduced[2]) if fixed_eta is None else float(fixed_eta)
        sse, evaluations = _evaluate_si(physical, eta, spectra)
        records.append(
            {
                "fixed_omega_p_substrate_cm_inv": fixed_substrate,
                "d_um": float(physical[0]),
                "omega_p_layer_cm_inv": float(physical[1]),
                "eta": eta,
                "mse_percent_squared": float(sse / sum(item.reflectance_percent.size for item in spectra)),
                "relative_mse_to_selected": float((sse / sum(item.reflectance_percent.size for item in spectra)) / reference_mse),
                "calibration": {
                    f"attachment_{item['attachment']}": {
                        "scale": float(item["scale"]),
                        "offset_percent": float(item["offset_percent"]),
                    }
                    for item in evaluations
                },
                "success": bool(result.success),
                "message": str(result.message),
            }
        )

    comparable = [item for item in records if item["relative_mse_to_selected"] <= 1.20]
    all_thickness = np.asarray([item["d_um"] for item in records], dtype=float)
    comparable_thickness = np.asarray([item["d_um"] for item in comparable], dtype=float)
    return {
        "purpose": "Conditional sensitivity to the silicon substrate Drude surrogate that reached the primary search bound.",
        "selected_model": mode,
        "fixed_values_cm_inv": list(fixed_values),
        "reoptimised_parameters": ["d_um", "omega_p_layer_cm_inv"]
        + ([] if fixed_eta is not None else ["eta"]),
        "comparison_rule": "Comparable means relative full-band MSE <= 1.20 times the selected fit; this is a transparent diagnostic threshold, not a confidence interval.",
        "reference_d_um": float(selected_parameters["d_um"]),
        "reference_mse_percent_squared": reference_mse,
        "records": records,
        "summary": {
            "all_trials_minimum_d_um": float(np.min(all_thickness)),
            "all_trials_maximum_d_um": float(np.max(all_thickness)),
            "comparable_trial_count": int(len(comparable)),
            "comparable_minimum_d_um": float(np.min(comparable_thickness)) if comparable_thickness.size else None,
            "comparable_maximum_d_um": float(np.max(comparable_thickness)) if comparable_thickness.size else None,
            "comparable_span_um": float(np.ptp(comparable_thickness)) if comparable_thickness.size else None,
            "comparable_maximum_relative_deviation": float(
                np.max(np.abs(comparable_thickness - selected_parameters["d_um"]))
                / selected_parameters["d_um"]
            )
            if comparable_thickness.size
            else None,
        },
    }


def _sic_predict(sigma: np.ndarray, angle_deg: float, physical: np.ndarray, eta: float) -> np.ndarray:
    d, gamma_l, gamma_t, wp_layer, gp_layer, wp_sub, gp_sub = np.asarray(physical, dtype=float)
    n_layer = sic_index(sigma, gamma_l, gamma_t, wp_layer, gp_layer)
    n_substrate = sic_index(sigma, gamma_l, gamma_t, wp_sub, gp_sub)
    return layered_reflectance_percent(sigma, n_layer, n_substrate, d, angle_deg, eta)


def _evaluate_sic(
    physical: np.ndarray,
    spectra: tuple[Spectrum, Spectrum],
    eta: float,
) -> tuple[float, list[dict[str, Any]]]:
    total_sse = 0.0
    evaluations: list[dict[str, Any]] = []
    for spectrum in spectra:
        physical_percent = _sic_predict(
            spectrum.sigma_cm_inv, spectrum.angle_deg, physical, eta
        )
        scale, offset, prediction = affine_calibration(
            physical_percent, spectrum.reflectance_percent
        )
        metrics = fit_metrics(spectrum.reflectance_percent, prediction)
        total_sse += metrics["sse_percent_squared"]
        evaluations.append(
            {
                "attachment": spectrum.attachment,
                "angle_deg": spectrum.angle_deg,
                "scale": scale,
                "offset_percent": offset,
                "metrics": metrics,
                "physical_percent": physical_percent,
                "prediction_percent": prediction,
            }
        )
    return total_sse, evaluations


def load_sic_q2_reference() -> tuple[np.ndarray, str]:
    sibling = ROOT.parent / "题目2_完整交付" / "results" / "q2_fit.json"
    if sibling.exists():
        payload = json.loads(sibling.read_text(encoding="utf-8"))
        params = payload["joint_fit"]["parameters"]
        vector = np.array([params[name] for name in SIC_PARAMETER_NAMES], dtype=float)
        return vector, str(sibling)
    return SIC_Q2_FALLBACK.copy(), "embedded_verified_q2_fallback"


def fit_sic_multibeam_correction(
    spectra: tuple[Spectrum, Spectrum],
    reference: np.ndarray,
    reference_source: str,
) -> dict[str, Any]:
    """Refit the exact Airy model near the verified Q2 solution."""

    # A correction check should not jump to a different phase alias.  The
    # neighbourhood is wide relative to the Q2 sensitivity interval but stays
    # on the same local fringe branch.
    local_bounds = (
        (6.8, 8.0),
        SIC_BOUNDS[1],
        SIC_BOUNDS[2],
        SIC_BOUNDS[3],
        SIC_BOUNDS[4],
        SIC_BOUNDS[5],
        SIC_BOUNDS[6],
    )

    def objective(unit: np.ndarray) -> float:
        physical = _decode(unit, local_bounds)
        sse, _ = _evaluate_sic(physical, spectra, eta=1.0)
        return sse / sum(item.reflectance_percent.size for item in spectra)

    starts = [
        reference,
        reference + np.array([0.10, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]),
        reference + np.array([-0.10, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]),
    ]
    candidates: list[tuple[np.ndarray, Any]] = []
    for start in starts:
        result = minimize(
            objective,
            _encode(start, local_bounds),
            method="L-BFGS-B",
            bounds=[(0.0, 1.0)] * len(local_bounds),
            options={"maxiter": 1600, "ftol": 1e-13, "gtol": 1e-10, "maxls": 70},
        )
        candidates.append((np.asarray(result.x, dtype=float), result))
    unit, result = min(candidates, key=lambda item: float(item[1].fun))
    fitted = _decode(unit, local_bounds)
    sse, evaluations = _evaluate_sic(fitted, spectra, eta=1.0)
    n_obs = sum(item.reflectance_percent.size for item in spectra)

    rho_by_attachment: dict[str, dict[str, float]] = {}
    for spectrum in spectra:
        n_layer = sic_index(
            spectrum.sigma_cm_inv, fitted[1], fitted[2], fitted[3], fitted[4]
        )
        n_substrate = sic_index(
            spectrum.sigma_cm_inv, fitted[1], fitted[2], fitted[5], fitted[6]
        )
        rho = round_trip_strength(
            spectrum.sigma_cm_inv,
            n_layer,
            n_substrate,
            fitted[0],
            spectrum.angle_deg,
        )
        rho_by_attachment[f"attachment_{spectrum.attachment}"] = {
            "median": float(np.median(rho)),
            "p95": float(np.quantile(rho, 0.95)),
            "maximum": float(np.max(rho)),
        }
    reference_sse, _ = _evaluate_sic(reference, spectra, eta=0.0)
    return {
        "model": "SiC_exact_Airy_eta_1",
        "purpose": "Question-3 correction check for omitted higher-order returns in Attachments 1 and 2.",
        "reference_q2_source": reference_source,
        "reference_q2_parameters": dict(zip(SIC_PARAMETER_NAMES, reference)),
        "local_search_bounds": {
            name: list(bound)
            for name, bound in zip(SIC_PARAMETER_NAMES, local_bounds)
        },
        "eta": 1.0,
        "parameters": dict(zip(SIC_PARAMETER_NAMES, fitted)),
        "calibration": {
            f"attachment_{item['attachment']}": {
                "scale": float(item["scale"]),
                "offset_percent": float(item["offset_percent"]),
            }
            for item in evaluations
        },
        "metrics": {
            f"attachment_{item['attachment']}": item["metrics"]
            for item in evaluations
        },
        "n_observations": int(n_obs),
        "full_mse_percent_squared": float(sse / n_obs),
        "reference_two_beam_mse_percent_squared": float(reference_sse / n_obs),
        "delta_thickness_um": float(fitted[0] - reference[0]),
        "relative_delta_thickness": float((fitted[0] - reference[0]) / reference[0]),
        "optimizer": {
            "method": "three deterministic local L-BFGS-B starts",
            "success": bool(result.success),
            "message": str(result.message),
            "nfev": int(getattr(result, "nfev", -1)),
        },
        "round_trip_strength": rho_by_attachment,
        "_evaluations": evaluations,
    }


def write_si_predictions(
    spectra: tuple[Spectrum, Spectrum],
    model_results: dict[str, dict[str, Any]],
    selected_model: str,
) -> None:
    path = RESULTS_DIR / "q3_si_predictions.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "attachment",
                "angle_deg",
                "wavenumber_cm_inv",
                "observed_percent",
                "two_beam_percent",
                "full_multibeam_percent",
                "partial_multibeam_percent",
                "selected_percent",
                "selected_residual_percent",
            ]
        )
        for spectrum in spectra:
            predictions: dict[str, np.ndarray] = {}
            for mode, result in model_results.items():
                evaluation = next(
                    item
                    for item in result["_evaluations"]
                    if item["attachment"] == spectrum.attachment
                )
                predictions[mode] = evaluation["prediction_percent"]
            selected = predictions[selected_model]
            for index in range(spectrum.sigma_cm_inv.size):
                writer.writerow(
                    [
                        spectrum.attachment,
                        spectrum.angle_deg,
                        spectrum.sigma_cm_inv[index],
                        spectrum.reflectance_percent[index],
                        predictions["two_beam"][index],
                        predictions["full_multibeam"][index],
                        predictions["partial_multibeam"][index],
                        selected[index],
                        selected[index] - spectrum.reflectance_percent[index],
                    ]
                )


def write_sic_predictions(
    spectra: tuple[Spectrum, Spectrum], correction: dict[str, Any]
) -> None:
    path = RESULTS_DIR / "q3_sic_correction_predictions.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "attachment",
                "angle_deg",
                "wavenumber_cm_inv",
                "observed_percent",
                "exact_multibeam_corrected_percent",
                "corrected_residual_percent",
            ]
        )
        for spectrum in spectra:
            evaluation = next(
                item
                for item in correction["_evaluations"]
                if item["attachment"] == spectrum.attachment
            )
            prediction = evaluation["prediction_percent"]
            for index in range(spectrum.sigma_cm_inv.size):
                writer.writerow(
                    [
                        spectrum.attachment,
                        spectrum.angle_deg,
                        spectrum.sigma_cm_inv[index],
                        spectrum.reflectance_percent[index],
                        prediction[index],
                        prediction[index] - spectrum.reflectance_percent[index],
                    ]
                )


def main() -> None:
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    all_spectra = {attachment: load_spectrum(attachment) for attachment in (1, 2, 3, 4)}
    audit = audit_inputs(all_spectra)

    si_source = (all_spectra[3], all_spectra[4])
    band_selection = select_si_band(si_source)
    selected_band = tuple(band_selection["selected"]["band_cm_inv"])
    si_spectra = tuple(_subset(item, selected_band) for item in si_source)
    initialisation = peak_initialisation(si_source, selected_band)
    initial_d = float(initialisation["mean_d_initial_um"])

    model_results = {
        mode: fit_si_model(mode, si_spectra, initial_d)
        for mode in ("two_beam", "full_multibeam", "partial_multibeam")
    }
    bic_two = model_results["two_beam"]["information_criteria"]["bic"]
    bic_full = model_results["full_multibeam"]["information_criteria"]["bic"]
    bic_partial = model_results["partial_multibeam"]["information_criteria"]["bic"]
    fitted_eta = model_results["partial_multibeam"]["parameters"]["eta"]
    delta_bic_partial = bic_two - bic_partial
    selected_model = (
        "partial_multibeam"
        if delta_bic_partial >= 10.0 and fitted_eta >= 0.05
        else "two_beam"
    )
    selected = model_results[selected_model]
    individual = [
        individual_si_fit(item, selected_model, selected["parameters"])
        for item in si_spectra
    ]
    substrate_sensitivity = si_substrate_sensitivity(si_spectra, selected)
    write_json(RESULTS_DIR / "q3_substrate_sensitivity.json", substrate_sensitivity)
    write_si_predictions(si_spectra, model_results, selected_model)

    sic_reference, reference_source = load_sic_q2_reference()
    sic_correction = fit_sic_multibeam_correction(
        (all_spectra[1], all_spectra[2]), sic_reference, reference_source
    )
    write_sic_predictions((all_spectra[1], all_spectra[2]), sic_correction)

    result_payload = {
        "problem": "2025 CUMCM B Question 3",
        "scope": {
            "user_request": "Complete Question 3 in an independent directory; preserve Question-2 and raw inputs.",
            "question_3_tasks": [
                "derive multi-beam necessary conditions and precision impact",
                "test Attachments 3 and 4 for Si multi-beam interference",
                "fit Si epitaxial thickness with the selected model",
                "check and correct Attachments 1 and 2 for omitted higher-order returns",
            ],
        },
        "input_audit": audit,
        "model_assumptions": {
            "geometry": "air--epitaxial layer--substrate",
            "phase": "delta = 4*pi*sigma*n_layer*cos(theta_layer)*d, sigma in cm^-1 and d in cm inside the phase",
            "polarisation": "unpolarised; arithmetic mean of s and p intensities",
            "silicon_background": {
                "model": "Sellmeier",
                "wavelength_um_validity": list(SI_SELLMEIER_WAVELENGTH_UM),
                "coefficients_B": list(SI_SELLMEIER_B),
                "coefficients_C": list(SI_SELLMEIER_C),
            },
            "silicon_free_carrier": "separate lossless Drude surrogates for epitaxial layer and substrate on the selected transparent band",
            "measurement_calibration": "per-angle bounded affine scale [0.25,1.75] and offset [-30,30] percent, analytically eliminated",
            "eta_definition": "eta=0 first-return truncation; eta=1 exact coherent Airy sum; fitted eta in [0,1] is a partial-coherence diagnostic",
            "sic_constants": SIC_CONSTANTS,
            "shared_si_parameters_across_angles": [
                "d_um",
                "omega_p_layer_cm_inv",
                "omega_p_substrate_cm_inv",
            ],
        },
        "multi_beam_conditions": {
            "successive_optical_path_difference": "Delta L = 2*n_layer*d*cos(theta_layer)",
            "phase_difference": "Delta phi = 4*pi*sigma*n_layer*d*cos(theta_layer)",
            "coherence_condition": "source/instrument coherence length must not be much shorter than Delta L",
            "amplitude_condition": "rho = |r10*r12*exp(i*Delta phi)| must be non-negligible; rho is the amplitude ratio of successive internal returns",
            "spectral_resolution_condition": "the instrument must resolve the phase variation rather than average it over many fringe periods",
            "necessary_not_sufficient_note": "These conditions are necessary for an observable multi-beam contribution; roughness, finite aperture, nonuniformity and averaging can still suppress it.",
            "precision_impact": "when rho is non-negligible, the Airy denominator changes line shape and extrema positions; fitting only the first return creates model bias in d, while the exact sum can recover the phase more efficiently.",
        },
        "si_band_selection": band_selection,
        "si_peak_initialisation": initialisation,
        "si_models": {
            mode: {key: value for key, value in result.items() if not key.startswith("_")}
            for mode, result in model_results.items()
        },
        "si_individual_angle_diagnostics": individual,
        "si_substrate_sensitivity": substrate_sensitivity,
        "si_model_selection": {
            "delta_bic_two_minus_full": float(bic_two - bic_full),
            "delta_bic_two_minus_partial": float(delta_bic_partial),
            "partial_fitted_eta": float(fitted_eta),
            "selection_rule": "Select partial multibeam only when Delta BIC(two minus partial)>=10 and fitted eta>=0.05; retain full eta=1 as the physical coherent-sum sensitivity model.",
            "selected_model": selected_model,
            "selected_thickness_um": float(selected["parameters"]["d_um"]),
        },
        "sic_correction": {key: value for key, value in sic_correction.items() if not key.startswith("_")},
        "prediction_files": {
            "si": "results/q3_si_predictions.csv",
            "sic_correction": "results/q3_sic_correction_predictions.csv",
        },
        "environment": {
            "python": sys.version,
            "numpy": np.__version__,
            "scipy": __import__("scipy").__version__,
            "openpyxl": __import__("openpyxl").__version__,
            "platform": platform.platform(),
        },
    }
    write_json(RESULTS_DIR / "q3_result.json", result_payload)

    print(
        json.dumps(
            {
                "selected_model": selected_model,
                "si_thickness_um": selected["parameters"]["d_um"],
                "si_eta": selected["parameters"]["eta"],
                "si_delta_bic_two_minus_partial": delta_bic_partial,
                "sic_reference_thickness_um": sic_reference[0],
                "sic_corrected_thickness_um": sic_correction["parameters"]["d_um"],
                "sic_delta_um": sic_correction["delta_thickness_um"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
